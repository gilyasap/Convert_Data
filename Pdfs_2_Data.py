import openpyxl
import pandas as pd
from pdf2image import convert_from_path
import os
from PIL import Image
import pytesseract
import shutil
import glob

current_directory = os.getcwd()
file_list = os.listdir(current_directory)
pdf_files = [file for file in file_list if file.lower().endswith('.pdf')]
for pdf_file in pdf_files:
    
    os.makedirs('images', exist_ok=True)
    images = convert_from_path(pdf_file)
    #From PDF to images
    for page_number, image in enumerate(images):
        image_filename = f"page_{page_number}.png"
        image_path = os.path.join('images', image_filename)
        image.save(image_path, "PNG")
    #From images to croped images
    input_folder = 'images'
    output_folder = 'Croped_images'
    os.makedirs(output_folder, exist_ok=True)
    top_crop = 65
    bottom_crop = 100

    for filename in os.listdir(input_folder):
        if filename.startswith('page_') and filename.endswith('.png'):
            page_number = int(filename.split('_')[1].split('.')[0])
            input_path = os.path.join(input_folder, filename)
            output_path = os.path.join(output_folder, filename)
            image = Image.open(input_path)
            width, height = image.size
            new_height = height - (top_crop + bottom_crop)
            cropped_image = image.crop((0, top_crop, width, height - bottom_crop))
            cropped_image.save(output_path)
    cwd = os.getcwd()
    dir=os.path.join(cwd,'Croped_images')
    png_files = glob.glob(os.path.join(dir, '*.png'))
    count = len(png_files)
    #croped images cut & convert to Text files            
    def cut_and_convert_images(image_folder, output_folder, part_widths):
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        all_text = []
        
        for i in range(2, count-1):
            image_name = f"page_{i}.png"
            image_path = os.path.join(image_folder, image_name)
            image = Image.open(image_path)
            width, height = image.size
            start = 0
            image_text = []
            for j, part_width in enumerate(part_widths):
                end = start + part_width
                part = image.crop((start, 0, end, height))
                text = pytesseract.image_to_string(part, config='--psm 3')
                image_text.append(text)
                start = end
            text_filename = f"page_{i}_text.txt"
            with open(os.path.join(output_folder, text_filename), "w") as text_file:
                text_file.write("\n".join(image_text))
            all_text.extend(image_text)
        with open(os.path.join(output_folder, "all_text.txt"), "w") as all_text_file:
            all_text_file.write("\n".join(all_text))
    image_folder = "Croped_images"
    output_folder = "Text_files"
    part_widths = [560, 550, 560]
    cut_and_convert_images(image_folder, output_folder, part_widths)#Calling Function

    #Clean the text files
    def clean_and_update_text_file(file_path, keywords, phrases_to_remove):
        with open(file_path, 'r+') as file:
            lines = file.readlines()
            file.seek(0)
            file.truncate(0)
            for line in lines:
                for keyword in keywords:
                    index = line.find(keyword)
                    if index != -1:
                        line = line[index:]
                for phrase in phrases_to_remove:
                    line = line.replace(phrase, '')
                    line = line.replace(": ",":")
                    line = line.replace("!","")
                    
                if line.startswith("Age"):
                    parts = line.split(" ", 1)
                    if len(parts) == 2:
                        file.write(parts[0] + '\n' + parts[1] + '\n')
                    else:
                        file.write(line)
                else:
                    file.write(line)
                                
    input_file_path = r'Text_files\\all_text.txt'
    keywords = ["XQK","KYJ","CVW"]
    phrases_to_remove = ["Photo", "is", "Available"]
    clean_and_update_text_file(input_file_path, keywords, phrases_to_remove)# Call the function

    data = pd.read_csv(r'Text_files\\all_text.txt', delimiter=',',encoding='Windows-1252',skip_blank_lines=True)
    name = os.path.splitext(pdf_file)[0] + f".xlsx"
    data.to_excel(name, index=False)
    data.to_excel(name, index=False)

    max_rows = 7000
    column_index_to_clean = 1
    excel_file = name
    df = pd.read_excel(excel_file, header=None)
    modified_data = []
    index = 0
    end_row = 7000
    names=['Name','Fathers Name','Husbands Name','House','Age','Gender']
    mask = df.iloc[:, 0].str.startswith(tuple(names))
    mask = mask.fillna(False)
    df = df[mask]
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)   

    workbook = openpyxl.load_workbook(name)
    sheet = workbook['Sheet1']
    #Moving rows from A to B
    for row in sheet.iter_rows(min_row=1, max_row=7500, min_col=1, max_col=1):
        cell_value = row[0].value
        cell_value = str(cell_value)
        if cell_value.startswith("Name"):
            row_index = row[0].row
            values_to_move = []
            for offset in range(1, 5):
                values_to_move.append(sheet.cell(row=row_index + offset, column=1).value)
            for col_index, value in enumerate(values_to_move):
                sheet.cell(row=row_index, column=col_index + 2, value=value)
            # Clear the original cells
            for offset in range(1, 5):
                sheet.cell(row=row_index + offset, column=1).value = None
    workbook.save(name)

    df = pd.read_excel(name, sheet_name='Sheet1')
    droped_cells = df.dropna()
    droped_cells.to_excel(name, sheet_name='Sheet1', index=False)

    wb = openpyxl.load_workbook(name)
    sheet1 = wb['Sheet1']
    for row in sheet1.iter_rows(min_row=1, min_col=4, max_row=1200, max_col=4):
        cell = row[0]
        cell_value = str(cell.value)
        if cell_value.startswith('Ag'):
            pass        
        else:
            sheet1.delete_rows(cell.row)    
    wb.save(name)

    wb = openpyxl.load_workbook(name)
    sheet1 = wb['Sheet1']
    for row in sheet1.iter_rows(min_row=1, min_col=5, max_row=1200, max_col=5):
        cell = row[0]
        cell_value = str(cell.value)
        if cell_value.startswith('Gen'):
            pass        
        else:
            sheet1.delete_rows(cell.row)    
    wb.save(name)

    wb = openpyxl.load_workbook(name)
    sheet1 = wb['Sheet1']
    rows_to_delete = []
    for row in sheet1.iter_rows(min_row=1, min_col=5, max_row=1200, max_col=5):
        cell = row[0]
        cell_value = str(cell.value)
        if cell_value.startswith('Name'):
            rows_to_delete.append(cell.row)
    for row_index in reversed(rows_to_delete):
        sheet1.delete_rows(row_index)
    wb.save(name)

    df = pd.read_excel(name, names=['Name', "Father's Name", 'House Number', 'Age', 'Gender'])
    df = pd.read_excel(name, header=None)
    df.columns = ["Name", "Father's Name", "House Number", "Age", "Gender"]
    df.insert(df.columns.get_loc("Father's Name") + 1, "Husband's Name", '')
    for index, row in df.iterrows():
        if row["Father's Name"].startswith("Husbands Name"):
            df.at[index, "Husband's Name"] = row["Father's Name"]
            df.at[index, "Father's Name"] = ''
    df['Name'] = df['Name'].str.replace('Name:', '')
    df['Name'] = df['Name'].str.replace('Name +', '')
    df['Name'] = df['Name'].str.replace('Name ?', '')
    df["Father's Name"] = df["Father's Name"].str.replace("Fathers Name:", '')
    df["Husband's Name"] = df["Husband's Name"].str.replace("Husbands Name:", '')
    df['House Number'] = df['House Number'].str.replace('House Number:', '')
    df['Age'] = df['Age'].str.replace('Age:', '')
    df['Gender'] = df['Gender'].str.replace('Gender:', '')
    
    shutil.rmtree('Croped_images')
    shutil.rmtree('images')
    shutil.rmtree('Text_files')
    name = os.path.splitext(pdf_file)[0] + f".xlsx"
    df.to_excel(name, index=False)

